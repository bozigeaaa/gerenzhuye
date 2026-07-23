import base64
import json
import os
import re
import tempfile
import time
from pathlib import Path
from urllib.parse import quote_plus, urljoin

import openpyxl
import requests


DOC_URL = "https://docs.qq.com/sheet/DTVdMaUJJTWd5WFha?tab=000008"
OUTPUT_PATH = Path("assets/shenzhen-food-data.json")
LOCAL_XLSX_TEXT = os.environ.get("SHENZHEN_FOOD_XLSX", "").strip()
LOCAL_XLSX = Path(LOCAL_XLSX_TEXT) if LOCAL_XLSX_TEXT else None
COOKIE = os.environ.get("TENCENT_DOCS_COOKIE", "").strip()

DISTRICT_FALLBACKS = {
    "宝安区": "宝安",
    "南山区": "南山",
    "福田区": "福田",
    "罗湖区": "罗湖",
    "龙岗区": "龙岗",
    "龙华区": "龙华",
    "盐田区": "盐田",
    "光明区": "光明",
    "坪山区": "坪山",
    "大鹏新区": "大鹏",
}

HEADER_ALIASES = {
    "name": ["店名", "餐厅", "店铺", "商家", "名称", "名字"],
    "address": ["地址", "位置", "定位"],
    "dishes": ["推荐菜", "推荐", "菜品", "吃什么", "招牌"],
    "average": ["人均", "价格", "价位", "消费"],
    "rating": ["评分", "打分", "推荐指数", "分数"],
    "note": ["备注", "评价", "说明", "补充"],
    "reason": ["避雷理由", "理由", "原因", "问题"],
}


def compact(value):
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_district(sheet_name):
    name = compact(sheet_name).replace(" ", "")
    return DISTRICT_FALLBACKS.get(name, name.replace("区", "") or "其他")


def find_header_map(rows):
    best_row = 0
    best_score = -1
    best_map = {}

    for row_index, row in enumerate(rows[:12]):
        mapping = {}
        score = 0
        for col_index, value in enumerate(row):
            header = compact(value)
            if not header:
                continue
            for field, aliases in HEADER_ALIASES.items():
                if any(alias in header for alias in aliases):
                    mapping[field] = col_index
                    score += 1
        if score > best_score:
            best_row = row_index
            best_score = score
            best_map = mapping

    return best_row, best_map


def value_from(row, mapping, field, fallback="待补充"):
    index = mapping.get(field)
    if index is None or index >= len(row):
        return fallback
    return compact(row[index]) or fallback


def parse_sheet(ws):
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    if not rows:
        return []

    header_row, mapping = find_header_map(rows)
    district = normalize_district(ws.title)
    is_avoid = "避雷" in ws.title
    records = []

    for row in rows[header_row + 1 :]:
        name = value_from(row, mapping, "name", "")
        if not name or name in {"店名", "餐厅", "商家", "名称"}:
            continue

        address = value_from(row, mapping, "address")
        dishes = value_from(row, mapping, "dishes")
        average = value_from(row, mapping, "average")
        rating = value_from(row, mapping, "rating")
        note = value_from(row, mapping, "reason" if is_avoid else "note")

        record = {
            "district": "避雷" if is_avoid else district,
            "name": name,
            "address": address,
            "dishes": dishes,
            "average": average,
            "rating": rating,
            "note": note,
        }
        if is_avoid:
            record["type"] = "avoid"
        records.append(record)

    return records


def parse_workbook(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    food = []
    avoid = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        records = parse_sheet(ws)
        if "避雷" in sheet_name:
            avoid.extend(records)
        else:
            food.extend(records)

    return food, avoid


def parse_basic_client_vars(html):
    match = re.search(r"window\.basicClientVars=JSON\.parse\(decodeURIComponent\(escape\(atob\('([^']+)'\)", html)
    if not match:
        raise RuntimeError("没有找到腾讯文档的 basicClientVars")
    raw = base64.b64decode(match.group(1))
    return json.loads(raw.decode("utf-8"))


def apply_cookie(session):
    if not COOKIE:
        return
    for pair in COOKIE.split(";"):
        if "=" not in pair:
            continue
        name, value = pair.split("=", 1)
        session.cookies.set(name.strip(), value.strip(), domain=".docs.qq.com")


def download_tencent_xlsx(target_path):
    session = requests.Session()
    apply_cookie(session)
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": DOC_URL,
    }
    page = session.get(DOC_URL, headers=headers, timeout=30)
    page.raise_for_status()
    client_vars = parse_basic_client_vars(page.text)
    pad_info = client_vars["docInfo"]["padInfo"]
    domain_id = pad_info.get("domainId", "300000000")
    pad_id = pad_info["padId"]
    doc_id = f"{domain_id}${pad_id}"
    tok = session.cookies.get("TOK") or ""

    export_url = f"https://docs.qq.com/v1/export/export_office?xsrf={quote_plus(tok)}"
    export_response = session.post(
        export_url,
        data={"docId": doc_id},
        headers={**headers, "Content-Type": "application/x-www-form-urlencoded"},
        timeout=30,
    )
    export_response.raise_for_status()
    export_data = export_response.json()
    if export_data.get("ret") not in (0, 200):
        raise RuntimeError(export_data.get("msg") or "腾讯文档导出失败")

    operation_id = export_data.get("operationId")
    if not operation_id:
        raise RuntimeError("腾讯文档没有返回导出任务 ID")

    for _ in range(60):
        progress_response = session.get(
            "https://docs.qq.com/v1/export/query_progress",
            params={"operationId": operation_id},
            headers=headers,
            timeout=30,
        )
        progress_response.raise_for_status()
        progress_data = progress_response.json()
        if progress_data.get("ret") not in (0, 200):
            raise RuntimeError(progress_data.get("msg") or "腾讯文档导出进度查询失败")

        if progress_data.get("progress") == 100 and progress_data.get("file_url"):
            file_url = urljoin("https://docs.qq.com", progress_data["file_url"])
            file_response = session.get(file_url, headers=headers, timeout=60)
            file_response.raise_for_status()
            target_path.write_bytes(file_response.content)
            return
        time.sleep(2)

    raise RuntimeError("腾讯文档导出超时")


def main():
    with tempfile.TemporaryDirectory() as temp_dir:
        xlsx_path = Path(temp_dir) / "shenzhen-food.xlsx"
        if LOCAL_XLSX and LOCAL_XLSX.exists():
            xlsx_path.write_bytes(LOCAL_XLSX.read_bytes())
        else:
            download_tencent_xlsx(xlsx_path)

        food, avoid = parse_workbook(xlsx_path)
        if not food and not avoid:
            raise RuntimeError("没有解析到任何店铺数据")

    payload = {
        "version": os.environ.get("SHENZHEN_FOOD_DATA_VERSION", "auto"),
        "updatedAt": time.strftime("%Y-%m-%d %H:%M:%S"),
        "source": "腾讯文档：深圳美食逛吃逛吃",
        "food": food,
        "avoid": avoid,
    }
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Synced {len(food)} food records and {len(avoid)} avoid records.")


if __name__ == "__main__":
    main()
